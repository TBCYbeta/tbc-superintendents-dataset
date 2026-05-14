"""Create confusion matrix from ground truth and predictions."""
import csv
import logging
from pathlib import Path
from typing import Optional

import click
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def normalize_id(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        if "." in text:
            as_float = float(text)
            if as_float.is_integer():
                return str(int(as_float))
        return str(int(text))
    except ValueError:
        return text


def row_key(row: dict) -> tuple[str, str, str]:
    name = str(row.get("name", "")).strip().lower()
    leaid1 = normalize_id(row.get("leaid1"))
    leaid2 = normalize_id(row.get("leaid2"))
    return name, leaid1, leaid2


def normalize_truth(value: object) -> Optional[str]:
    text = str(value).strip().lower()
    if text in {"1", "same", "s", "true"}:
        return "same"
    if text in {"0", "different", "d", "false"}:
        return "different"
    if text in {"?", "unknown", ""}:
        return "unknown"
    return None


def normalize_prediction(value: object) -> Optional[str]:
    text = str(value).strip().lower()
    if text in {"same", "different"}:
        return text
    return None


def load_ground_truth(ground_truth_file: Path) -> dict[tuple[str, str, str], str]:
    """Load ground truth from Excel or CSV. Returns dict keyed by (name, leaid1, leaid2)."""
    ground_truth: dict[tuple[str, str, str], str] = {}

    if ground_truth_file.suffix.lower() == ".xlsx":
        wb = load_workbook(ground_truth_file, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {h: row[idx] for h, idx in header_map.items() if idx < len(row)}
            truth = normalize_truth(row_dict.get("prediction"))
            if truth is None:
                continue
            key = row_key(row_dict)
            ground_truth[key] = truth
        wb.close()
    else:
        with open(ground_truth_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                truth = normalize_truth(row.get("prediction"))
                if truth is None:
                    continue
                key = row_key(row)
                ground_truth[key] = truth

    return ground_truth


def load_predictions(predictions_csv: Path) -> dict[tuple[str, str, str], str]:
    predictions: dict[tuple[str, str, str], str] = {}
    with open(predictions_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pred = normalize_prediction(row.get("prediction"))
            if pred is None:
                continue
            predictions[row_key(row)] = pred
    return predictions


def create_confusion_matrix(
    ground_truth: dict[tuple[str, str, str], str],
    predictions: dict[tuple[str, str, str], str],
) -> dict:
    matrix = {
        ("same", "same"): 0,
        ("same", "different"): 0,
        ("different", "same"): 0,
        ("different", "different"): 0,
    }
    missing_ground_truth = 0
    missing_prediction = 0
    unknown_ground_truth = 0

    for key, pred in predictions.items():
        if key not in ground_truth:
            missing_ground_truth += 1
            continue
        truth = ground_truth[key]
        if truth == "unknown":
            unknown_ground_truth += 1
            continue
        matrix[(truth, pred)] += 1

    for key, truth in ground_truth.items():
        if truth != "unknown" and key not in predictions:
            missing_prediction += 1

    tp = matrix[("same", "same")]
    tn = matrix[("different", "different")]
    fp = matrix[("different", "same")]
    fn = matrix[("same", "different")]
    total = tp + tn + fp + fn
    accuracy = (tp + tn) / total if total else 0.0
    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0

    return {
        "matrix": matrix,
        "metrics": {
            "accuracy": accuracy,
            "precision": precision,
            "recall": recall,
            "f1": f1,
        },
        "skipped": {
            "missing_ground_truth": missing_ground_truth,
            "missing_prediction": missing_prediction,
            "unknown_ground_truth": unknown_ground_truth,
        },
    }


@click.command()
@click.argument("predictions_csv", type=click.Path(exists=True, path_type=Path))
@click.argument("ground_truth_file", type=click.Path(exists=True, path_type=Path))
@click.option("-o", "--output-file", type=click.Path(path_type=Path))
def main(predictions_csv: Path, ground_truth_file: Path, output_file: Optional[Path]) -> None:
    logger.info("Loading predictions from %s", predictions_csv)
    predictions = load_predictions(predictions_csv)
    logger.info("Loaded %s predictions (excluding errors)", len(predictions))

    logger.info("Loading ground truth from %s", ground_truth_file)
    ground_truth = load_ground_truth(ground_truth_file)
    logger.info("Loaded %s ground truth labels", len(ground_truth))

    results = create_confusion_matrix(ground_truth, predictions)
    matrix = results["matrix"]
    metrics = results["metrics"]
    skipped = results["skipped"]

    output = []
    output.append("=" * 60)
    output.append("CONFUSION MATRIX (truth vs prediction)")
    output.append("=" * 60)
    output.append("                Pred: same    Pred: different")
    output.append(f"Truth: same      {matrix[('same', 'same')]:>6}    {matrix[('same', 'different')]:>6}")
    output.append(
        f"Truth: different {matrix[('different', 'same')]:>6}    "
        f"{matrix[('different', 'different')]:>6}"
    )
    output.append("")
    output.append("=" * 60)
    output.append("METRICS")
    output.append("=" * 60)
    output.append(f"Accuracy:     {metrics['accuracy']:.1%}")
    output.append(f"Precision:    {metrics['precision']:.1%}")
    output.append(f"Recall:       {metrics['recall']:.1%}")
    output.append(f"F1:           {metrics['f1']:.3f}")
    output.append("")
    output.append("=" * 60)
    output.append("SKIPPED CASES")
    output.append("=" * 60)
    output.append(f"Missing ground truth:      {skipped['missing_ground_truth']}")
    output.append(f"Missing predictions:       {skipped['missing_prediction']}")
    output.append(f"Unknown ground truth (?):  {skipped['unknown_ground_truth']}")
    output.append("=" * 60)

    result_text = "\n".join(output)
    print(result_text)

    if output_file:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(result_text)
        logger.info("Results saved to %s", output_file)


if __name__ == "__main__":
    main()
